from openpyxl import load_workbook
wb = load_workbook(r"C:\Users\USER\Desktop\홈페이지갱신_미리보기.xlsx")
s = wb.active
rows = list(s.iter_rows(values_only=True))
header = rows[0]
print("샘플 (변경 대상 188건 중 무작위 12건):")
print()
import random
random.seed(42)
sample = random.sample(rows[1:], 12)
for r in sample:
    cn = r[2]
    cur_p, new_p = r[4], r[5]
    cur_e, new_e = r[6], r[7]
    cur_c, new_c = r[8], r[9]
    fields = r[10]
    print(f"  [{cn}]")
    if "phone" in (fields or ""):
        print(f"    phone: '{cur_p}' → '{new_p}'")
    if "email" in (fields or ""):
        print(f"    email: '{cur_e}' → '{new_e}'")
    if "contactName" in (fields or ""):
        print(f"    contact: '{cur_c}' → '{new_c}'")
    print()
