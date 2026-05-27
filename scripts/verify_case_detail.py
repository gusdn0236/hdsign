"""정리본 엑셀이 자재 디테일까지 들어있는지 검증.

특정 거래 사례 (큰 금액 + 'XX 현장') 의 모든 행을 거래처+발행일자로 묶어 확인.
"""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook

NAMESEO_FILES = [
    Path(r"C:\Users\USER\Downloads\files\25년도매출_개인_정리본.xlsx"),
    Path(r"C:\Users\USER\Downloads\files\25년도매출_주식회사_정리본.xlsx"),
]


def all_rows():
    for path in NAMESEO_FILES:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb["거래명세"]
        header = None
        for raw in ws.iter_rows(values_only=True):
            if header is None:
                header = list(raw)
                continue
            d = dict(zip(header, raw))
            d["_src"] = path.stem
            yield d


def show(filter_fn, label: str) -> None:
    print(f"\n========== {label} ==========")
    matched = [r for r in all_rows() if filter_fn(r)]
    print(f"matched: {len(matched)} rows")
    for r in matched:
        cat = (r.get("대분류") or "?")[:14]
        item = (r.get("품목") or "")[:50]
        supply = r.get("공급가액")
        try:
            s = f"₩{int(supply):>11,}"
        except Exception:
            s = f"{supply!s:>12}"
        date = str(r.get("발행일자") or "")[:10]
        client = (r.get("거래처명") or "")[:25]
        print(f"  {date} | {client:25s} | [{cat:14s}] | {s} | {item}")


def main() -> int:
    # 진성커뮤니티 가장 큰 case
    show(
        lambda r: r.get("거래처명") == "(주)진성커뮤니티" and str(r.get("발행일자") or "")[:10] == "2025-07-30",
        "진성커뮤니티 2025-07-30 (양주 지웰엘리움 ₩20.8M)",
    )
    # 진성커뮤니티 1월 첫 case
    show(
        lambda r: r.get("거래처명") == "(주)진성커뮤니티" and str(r.get("발행일자") or "")[:10] == "2025-01-03",
        "진성커뮤니티 2025-01-03 (잭니클라우스 ₩120k)",
    )
    # 갈바후렘 1.6T (자재 명시된 케이스)
    show(
        lambda r: r.get("거래처명") == "라이트너스" and str(r.get("발행일자") or "")[:10] == "2025-04-30",
        "라이트너스 2025-04-30 (갈바후렘1.6T ₩6.14M)",
    )
    # 디자인펜의 1.5M+ 큰 케이스 한 개
    show(
        lambda r: r.get("거래처명") == "세종그래픽디자인" and str(r.get("발행일자") or "")[:10] == "2025-01-08",
        "세종그래픽디자인 2025-01-08 (남영역 엘리베이터 ₩1.5M)",
    )
    # 디온에이싸인 (개인) 큰 케이스 한 개 dryrun 결과 보고 임의
    show(
        lambda r: r.get("거래처명") == "(주)다이노에프에스 서울사무소" and str(r.get("발행일자") or "")[:10] == "2025-03-20",
        "(주)다이노에프에스 서울사무소 2025-03-20",
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
