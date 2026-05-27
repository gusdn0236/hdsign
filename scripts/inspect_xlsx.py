"""엑셀 구조 dump — 시트명, 시트별 dims, 첫 N행."""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook


def dump(xlsx_path: Path, rows: int = 8) -> None:
    print(f"\n========== {xlsx_path.name} ==========")
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        print(f"\n--- sheet: '{ws.title}' ---")
        # read_only 모드에선 max_row 불안정 → 직접 카운트
        all_rows = list(ws.iter_rows(values_only=True))
        print(f"rows: {len(all_rows)}, cols(max from first row): "
              f"{len(all_rows[0]) if all_rows else 0}")
        for i, row in enumerate(all_rows[:rows], 1):
            cells = []
            for v in row:
                if v is None:
                    cells.append("·")
                elif isinstance(v, str):
                    cells.append(repr(v[:30]) if len(v) > 30 else repr(v))
                else:
                    cells.append(str(v))
            print(f"  r{i}: " + " | ".join(cells))


if __name__ == "__main__":
    for path in sys.argv[1:]:
        dump(Path(path))
