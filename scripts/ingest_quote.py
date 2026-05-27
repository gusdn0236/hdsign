"""
견적 ingest 파서 — 첫 이지폼 .xlsx 한 개 보고 채우는 골격.

사용:
    py -3 scripts\ingest_quote.py <inbox_folder>

출력:
    parsed.json (같은 폴더에) — {
      "items": [{"name": str, "spec": str, "qty": int, "unit_price": int, "amount": int, ...}],
      "subtotal": int, "tax": int, "total": int,
      "client_name": str | None,
      "issued_date": str | None
    }

이걸 Claude 가 읽어서 case 페이지의 knownCostTotal 표·finalPrice 채움.

TODO (첫 .xlsx 받으면 채울 영역):
    - HEADER_ROW / DATA_START_ROW: 이지폼은 양식이 고정인데 PC마다 시트 한 칸 밀린 경우 있음
    - COL_MAP: {col_letter or index: field_name} — 품목명/규격/수량/단가/금액 컬럼이 몇 번째인지
    - 합계/세금/총액 셀 위치
    - 거래처명 셀 (보통 상단)
    - 발행일 셀
"""

from __future__ import annotations
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl 필요: py -3 -m pip install --user openpyxl", file=sys.stderr)
    sys.exit(2)


# ↓↓↓ 첫 .xlsx 받으면 여기 채움 ↓↓↓
HEADER_ROW = 0          # 헤더가 N번째 행 (1-indexed) — 0 = 미설정
DATA_START_ROW = 0      # 품목 데이터 시작 행
COL_MAP: dict[str, int] = {
    # "name": 2,         # B열
    # "spec": 3,
    # "qty": 4,
    # "unit_price": 5,
    # "amount": 6,
}
CLIENT_NAME_CELL = ""   # 예: "C3"
ISSUED_DATE_CELL = ""   # 예: "G3"
TOTAL_CELL = ""         # 예: "G50" — 시트마다 다르면 마지막 비어있지 않은 amount 합산으로 폴백
# ↑↑↑ 첫 .xlsx 받으면 여기 채움 ↑↑↑


def parse_xlsx(xlsx_path: Path) -> dict:
    if HEADER_ROW == 0:
        raise RuntimeError(
            "파서 미설정. 첫 .xlsx 1개 보고 HEADER_ROW / COL_MAP 등을 채우세요.\n"
            "  - 시트 이름: " + str([s.title for s in load_workbook(xlsx_path).worksheets])
        )
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    items: list[dict] = []
    r = DATA_START_ROW
    while True:
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            break
        item = {}
        for field, col in COL_MAP.items():
            item[field] = row[col - 1]
        if item.get("name"):
            items.append(item)
        r += 1
    out = {
        "sheet": ws.title,
        "items": items,
        "row_count": len(items),
    }
    if CLIENT_NAME_CELL:
        out["client_name"] = ws[CLIENT_NAME_CELL].value
    if ISSUED_DATE_CELL:
        v = ws[ISSUED_DATE_CELL].value
        out["issued_date"] = str(v) if v else None
    if TOTAL_CELL:
        out["total"] = ws[TOTAL_CELL].value
    else:
        out["total"] = sum(
            (it.get("amount") or 0) for it in items if isinstance(it.get("amount"), (int, float))
        )
    return out


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("usage: ingest_quote.py <inbox_folder>", file=sys.stderr)
        return 2
    folder = Path(argv[1])
    if not folder.is_dir():
        print(f"not a directory: {folder}", file=sys.stderr)
        return 2
    xlsx_candidates = [
        p for p in folder.iterdir()
        if p.suffix.lower() == ".xlsx" and not p.name.startswith("~$")
    ]
    if not xlsx_candidates:
        print(f"no .xlsx in {folder}", file=sys.stderr)
        return 1
    if len(xlsx_candidates) > 1:
        print(f"여러 xlsx 발견 — 첫 번째 사용: {[p.name for p in xlsx_candidates]}", file=sys.stderr)
    xlsx_path = xlsx_candidates[0]
    parsed = parse_xlsx(xlsx_path)
    out_path = folder / "parsed.json"
    out_path.write_text(
        json.dumps(parsed, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"wrote {out_path}")
    print(f"  items: {parsed['row_count']}, total: {parsed.get('total')}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
