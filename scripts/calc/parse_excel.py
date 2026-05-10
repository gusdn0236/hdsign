"""
잔넬스카시단가표.xlsx → prices_excel.json

Reads the customer-supplied Excel and emits JSON in the same shape as
prices_baseline.json so the diff engine can compare cell-by-cell.

The Excel uses several non-machine-friendly conventions:
  - Some cells contain text like "영문기본 53,000" meaning "English base price for
    smallest sizes". The 'base' price applies to multiple sizes — we use the
    baseline's price-equality grouping to know which sizes share a base.
  - Numbers may contain comma separators ("53,000") or live as ints ("42000").
  - Sizes are encoded in cm (multiply by 10 for mm).
  - Merged cells span multiple rows/cols — we expand them so every (r,c)
    reads the top-left value.
  - Two side-by-side tables in the epoxy sheet (갈바 left, 스텐 right).

Usage:
    py scripts/calc/parse_excel.py [path/to/xlsx]

Reads:  default = C:\\Users\\USER\\Documents\\카카오톡 받은 파일\\잔넬스카시단가표.xlsx
        baseline = frontend/src/data/calc/prices_baseline.json
Writes: frontend/src/data/calc/prices_excel.json
"""
import json
import re
import sys
import io
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = Path(r"C:\Users\USER\Documents\카카오톡 받은 파일\잔넬스카시단가표.xlsx")
BASELINE_PATH = ROOT / "frontend" / "src" / "data" / "calc" / "prices_baseline.json"
OUT = ROOT / "frontend" / "src" / "data" / "calc" / "prices_excel.json"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def expand_merges(ws) -> dict:
    """Return {(row, col) → value}. Merged-cell members inherit top-left value."""
    grid = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            grid[(r, c)] = ws.cell(row=r, column=c).value
    for mr in ws.merged_cells.ranges:
        top_left = ws.cell(row=mr.min_row, column=mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                grid[(r, c)] = top_left
    return grid


def parse_number(s) -> int | None:
    """Coerce numeric strings ('53,000', '42000') and ints/floats → int. None on fail."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return int(s)
    if isinstance(s, str):
        s = s.strip().replace(",", "").replace("\n", "").replace(" ", "")
        if not s:
            return None
        # Strip leading text, keep only digits
        m = re.search(r"\d{3,}", s)  # at least 3 digits to avoid false positives
        if m:
            return int(m.group(0))
    return None


CHANNEL_LANG_PATTERNS = [
    # All four patterns are 'base': they convey the price for that lang's base
    # zone (smallest sizes). Excel uses "영문기본 N" most of the time but stenOsai
    # uses just "영문 N" / "한글 N" with the same meaning. Always propagate to
    # baseline's base group so the diff doesn't report a flood of false 'missing'.
    (re.compile(r"영문기본\s*([\d,]+)"),  "eng", True),
    (re.compile(r"한글기본\s*([\d,]+)"),  "kor", True),
    (re.compile(r"영문\s+([\d,]+)"),     "eng", True),
    (re.compile(r"한글\s+([\d,]+)"),     "kor", True),
]


def parse_channel_text_cell(text: str):
    """Return (lang, price, is_base) or None."""
    if not isinstance(text, str):
        return None
    flat = text.replace("\n", " ")
    for pat, lang, is_base in CHANNEL_LANG_PATTERNS:
        m = pat.search(flat)
        if m:
            n = int(m.group(1).replace(",", ""))
            return (lang, n, is_base)
    return None


def find_base_group(prices: dict) -> list[str]:
    """For a {sizeStr: price} map, return the leading group of consecutive sizes that share the smallest price."""
    if not prices:
        return []
    sizes_sorted = sorted(prices.keys(), key=int)
    base_price = prices[sizes_sorted[0]]
    out = []
    for s in sizes_sorted:
        if prices[s] == base_price:
            out.append(s)
        else:
            break
    return out


# ---------------------------------------------------------------------------
# Sheet 0: 잔넬
# ---------------------------------------------------------------------------
# Column layout (1-indexed): C01=size(cm), C02..C11=10 types in baseline order
CHANNEL_COL_TYPES = [
    # (col, baseline_type_key, baseline_label)
    (2,  "galvaBackEng", "갈바후광영문"),
    (3,  "galvaBackKor", "갈바후광한글"),
    (4,  "galvaOsai",    "갈바오사이"),
    (5,  "galvaCap",     "갈바캡잔넬"),
    (6,  "ilcheType",    "일체형잔넬"),
    (7,  "takaType",     "타카잔넬"),
    (8,  "stenAlumCap",  "스텐알미늄캡"),
    (9,  "stenOsai",     "스텐오사이"),
    (10, "stenBack",     "스텐후광"),
    (11, "goldSten",     "골드스텐"),
]
CHANNEL_DATA_ROWS = range(3, 30)  # R03 (size 20cm = 200mm) .. R29 (size 200cm = 2000mm)


def parse_channel(ws, baseline_channel) -> dict:
    grid = expand_merges(ws)
    baseline_types_by_key = {t["key"]: t for t in baseline_channel["types"]}

    types_out = []
    for col, key, label in CHANNEL_COL_TYPES:
        bt = baseline_types_by_key[key]
        needs_lang = bt["needsLang"]

        if needs_lang:
            base_group = {
                "eng": find_base_group(bt["pricesByLang"]["eng"]),
                "kor": find_base_group(bt["pricesByLang"]["kor"]),
            }
            prices_eng: dict[str, int] = {}
            prices_kor: dict[str, int] = {}

            for r in CHANNEL_DATA_ROWS:
                size_cm = grid.get((r, 1))
                if not isinstance(size_cm, (int, float)):
                    continue
                size_mm = int(size_cm) * 10
                size_str = str(size_mm)

                cell = grid.get((r, col))
                if cell is None or cell == "":
                    continue

                # Numeric cell → both langs share at this size
                num = parse_number(cell) if not isinstance(cell, str) else None
                if num is not None:
                    prices_eng[size_str] = num
                    prices_kor[size_str] = num
                    continue

                # Text cell → parse "영문기본 N" / "한글기본 N" / "영문 N" / "한글 N"
                if isinstance(cell, str):
                    parsed = parse_channel_text_cell(cell)
                    if parsed:
                        lang, price, is_base = parsed
                        target = prices_eng if lang == "eng" else prices_kor
                        if is_base:
                            # Propagate to baseline's known base group for this lang
                            for bs in base_group[lang]:
                                target[bs] = price
                        else:
                            target[size_str] = price
                    else:
                        # Couldn't parse — try just extracting a number as fallback
                        n = parse_number(cell)
                        if n is not None:
                            prices_eng[size_str] = n
                            prices_kor[size_str] = n

            types_out.append({
                "key": key,
                "label": label,
                "needsLang": True,
                "pricesByLang": {
                    "kor": dict(sorted(prices_kor.items(), key=lambda kv: int(kv[0]))),
                    "eng": dict(sorted(prices_eng.items(), key=lambda kv: int(kv[0]))),
                },
            })
        else:
            prices: dict[str, int] = {}
            for r in CHANNEL_DATA_ROWS:
                size_cm = grid.get((r, 1))
                if not isinstance(size_cm, (int, float)):
                    continue
                size_mm = int(size_cm) * 10
                cell = grid.get((r, col))
                num = parse_number(cell)
                if num is not None:
                    prices[str(size_mm)] = num
            types_out.append({
                "key": key,
                "label": label,
                "needsLang": False,
                "prices": prices,
            })

    return {
        "label": baseline_channel["label"],
        "sheetName": ws.title,
        "sizeAxis": baseline_channel["sizeAxis"],
        "types": types_out,
    }


# ---------------------------------------------------------------------------
# Sheet 1: 스카시 → gomu
# ---------------------------------------------------------------------------
GOMU_COL_THICKNESS = [
    (2, "10T"),
    (3, "10T-금은색"),
    (4, "20,30T"),
    (5, "20,30T-금은색"),
    (6, "50T"),
    (7, "50T-금은색"),
]


def gomu_band_for_size(size_mm: int) -> str:
    """Map an Excel point size (mm) → baseline band label. Mirrors extract_baseline.gomu_band logic."""
    if size_mm <= 149:
        return "~149"
    if size_mm <= 999:
        # JS: ceil((gomuH-149)/50) gives a row index; map to corresponding band
        # row 1 = 150~199, row 2 = 200~249, ..., row 17 = 950~999
        row_idx = (size_mm - 150) // 50 + 1
        low = 150 + (row_idx - 1) * 50
        return f"{low}~{low + 49}"
    # 1000mm and above land in 100mm bands; row 18 = 1000~1099
    row_idx = 18 + (size_mm - 1000) // 100
    low = 1000 + (row_idx - 18) * 100
    return f"{low}~{low + 99}"


GOMU_DATA_ROWS = range(4, 33)  # R04 (10cm = 100mm) .. R32 (200cm = 2000mm), 29 rows


def parse_gomu(ws, baseline_gomu) -> dict:
    grid = expand_merges(ws)
    structured = {tk: {} for _, tk in GOMU_COL_THICKNESS}

    for r in GOMU_DATA_ROWS:
        size_cm = grid.get((r, 1))
        if not isinstance(size_cm, (int, float)):
            continue
        size_mm = int(size_cm) * 10
        band = gomu_band_for_size(size_mm)

        for col, tk in GOMU_COL_THICKNESS:
            n = parse_number(grid.get((r, col)))
            if n is not None:
                structured[tk][band] = n

    return {
        "label": baseline_gomu["label"],
        "sheetName": ws.title,
        "axes": baseline_gomu["axes"],
        "_heightBandRule": baseline_gomu["_heightBandRule"],
        "prices": structured,
    }


# ---------------------------------------------------------------------------
# Sheet 2: 아크릴.포맥스 → acryl
# ---------------------------------------------------------------------------
# Cols 2-15 = 7 thicknesses × 2 langs (영, 한)
ACRYL_THICKNESSES = ["2T", "3T", "5T", "8T", "10T", "15T", "20T"]
ACRYL_TEXT_TYPES = ["영문", "한글"]


_BAND_RE = re.compile(r"^(~\d+|\d+~\d+|\d+)$")


def normalize_band_label(s) -> str | None:
    """Excel uses both '~' and '-' as range delimiters; '301-310' → '301~310'.
    Returns None for anything that doesn't look like a clean band label
    (e.g., footer notes, phone numbers, blank cells, plain numbers we don't want)."""
    if not isinstance(s, str):
        return None
    s = s.strip().replace("mm", "").replace(" ", "")
    if not s:
        return None
    s = s.replace("-", "~")
    if not _BAND_RE.match(s):
        return None
    return s


def parse_acryl(ws, baseline_acryl) -> dict:
    grid = expand_merges(ws)
    structured = {tk: {tt: {} for tt in ACRYL_TEXT_TYPES} for tk in ACRYL_THICKNESSES}

    # Walk rows starting at R05 until size column 1 stops being a band label
    for r in range(5, ws.max_row + 1):
        band = normalize_band_label(grid.get((r, 1)))
        if band is None:
            continue
        # baseline uses '~30' (no leading), '31~40', etc.
        for t_idx, tk in enumerate(ACRYL_THICKNESSES):
            for tt_idx, tt in enumerate(ACRYL_TEXT_TYPES):
                col = 2 + t_idx * 2 + tt_idx
                n = parse_number(grid.get((r, col)))
                if n is not None:
                    structured[tk][tt][band] = n

    return {
        "label": baseline_acryl["label"],
        "sheetName": ws.title,
        "axes": baseline_acryl["axes"],
        "_heightBandRule": baseline_acryl["_heightBandRule"],
        "prices": structured,
    }


# ---------------------------------------------------------------------------
# Sheet 3: 금은경 → goldSilver (NEW; raw extract — Phase 4 will canonicalize)
# ---------------------------------------------------------------------------
# Layout (from xlsx_dump):
#   R02 thickness header: 2T, 2T, 3T, 3T, 5T, 5T, 8T, 8T, 10T, 8T, 8T, 10T, 10T, 15T, 15T, 20T, 20T
#   R03 lang header:      영문, 한글, 영문, 한글, 영문, 한글, 영문, 한글, 영문, 영문, 한글, 영문, 한글, 영문, 한글, 영문, 한글
#   R04+ data
# The two halves likely correspond to 금경 (gold mirror, cols 2-10) vs 은경 (silver mirror, cols 11-18).
# Col 10 has only 영문 for 10T (가운데 단독). Phase 4 will confirm with user.
GOLD_SILVER_COLUMNS = [
    # (col, material_key, thickness, textType)
    (2,  "gold",   "2T",  "영문"),
    (3,  "gold",   "2T",  "한글"),
    (4,  "gold",   "3T",  "영문"),
    (5,  "gold",   "3T",  "한글"),
    (6,  "gold",   "5T",  "영문"),
    (7,  "gold",   "5T",  "한글"),
    (8,  "gold",   "8T",  "영문"),
    (9,  "gold",   "8T",  "한글"),
    (10, "gold",   "10T", "영문"),
    (11, "silver", "8T",  "영문"),
    (12, "silver", "8T",  "한글"),
    (13, "silver", "10T", "영문"),
    (14, "silver", "10T", "한글"),
    (15, "silver", "15T", "영문"),
    (16, "silver", "15T", "한글"),
    (17, "silver", "20T", "영문"),
    (18, "silver", "20T", "한글"),
]


def parse_gold_silver(ws) -> dict:
    grid = expand_merges(ws)
    # Use {material → thickness → textType → band → price}
    structured: dict = {"gold": {}, "silver": {}}

    bands_seen = []
    for r in range(4, ws.max_row + 1):
        band = normalize_band_label(grid.get((r, 1)))
        if band is None:
            continue
        bands_seen.append(band)
        for col, mat, tk, tt in GOLD_SILVER_COLUMNS:
            n = parse_number(grid.get((r, col)))
            if n is None:
                continue
            structured.setdefault(mat, {}).setdefault(tk, {}).setdefault(tt, {})[band] = n

    return {
        "label": "금은경 (금경/은경 아크릴)",
        "sheetName": ws.title,
        "_status": "Phase 2 raw extract — Phase 4 에서 사용자와 구조 확정 필요",
        "axes": {
            "material": [
                {"key": "gold",   "label": "금경"},
                {"key": "silver", "label": "은경"},
            ],
            "thicknessByMaterial": {
                "gold":   ["2T", "3T", "5T", "8T", "10T"],
                "silver": ["8T", "10T", "15T", "20T"],
            },
            "textType": ["영문", "한글"],
            "heightBands": bands_seen,
        },
        "prices": structured,
    }


# ---------------------------------------------------------------------------
# Sheet 4: 에폭시잔넬고딕체 → epoxy
# ---------------------------------------------------------------------------
# Two side-by-side tables:
#   Cols 1-7: size_band | textType | 갈바 stroke 30/50/70/90/110
#   Col 8: separator (skip)
#   Cols 9-15: size_band | textType | 스텐 stroke 30/50/70/90/110
# Rows 5-30: data, paired (한글 row, 영문숫자 row) per size band
EPOXY_STROKES = [30, 50, 70, 90, 110]
EPOXY_TEXTTYPE_KEY = {
    "한글": "korean",
    "영문숫자": "englishNumber",
    "영문/숫자": "englishNumber",
    "영문 숫자": "englishNumber",
}


def parse_epoxy(ws, baseline_epoxy) -> dict:
    grid = expand_merges(ws)
    structured: dict = {
        "galvalume": {"korean": {}, "englishNumber": {}},
        "stainless": {"korean": {}, "englishNumber": {}},
    }

    sections = [
        # (material_key, size_col, textType_col, price_start_col)
        ("galvalume", 1, 2, 3),
        ("stainless", 9, 10, 11),
    ]

    for r in range(5, ws.max_row + 1):
        for material, size_col, tt_col, price_start_col in sections:
            band_raw = grid.get((r, size_col))
            band = normalize_band_label(band_raw)
            if band is None or not band.startswith("~"):
                continue
            # Strip the ~ and parse the size number
            try:
                size_mm = int(band.lstrip("~").strip())
            except ValueError:
                continue

            tt_raw = grid.get((r, tt_col))
            if not isinstance(tt_raw, str):
                continue
            tt_key = EPOXY_TEXTTYPE_KEY.get(tt_raw.strip())
            if tt_key is None:
                continue

            for stroke_idx, stroke in enumerate(EPOXY_STROKES):
                n = parse_number(grid.get((r, price_start_col + stroke_idx)))
                if n is not None:
                    structured[material][tt_key].setdefault(str(size_mm), {})[str(stroke)] = n

    return {
        "label": baseline_epoxy["label"],
        "sheetName": ws.title,
        "axes": baseline_epoxy["axes"],
        "prices": structured,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        print(f"ERROR: xlsx not found: {xlsx_path}")
        sys.exit(1)

    baseline = json.loads(BASELINE_PATH.read_text(encoding="utf-8"))
    bcalc = baseline["calculators"]

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheets = {ws.title.strip(): ws for ws in wb.worksheets}

    # Match sheet names by leading keyword (xlsx uses trailing whitespace and Korean text)
    def find_sheet(*keywords):
        for title, ws in sheets.items():
            if any(kw in title for kw in keywords):
                return ws
        return None

    out = {
        "_meta": {
            "version": "excel-parsed",
            "extractedFrom": str(xlsx_path),
            "extractor": "scripts/calc/parse_excel.py",
            "doc": (
                "엑셀 → JSON 변환 결과. baseline 구조와 동일한 shape. "
                "값은 엑셀 그대로(오타·누락 가능). diff 엔진(Phase 3)에서 baseline과 비교 후 "
                "사용자 승인 거쳐 prices.json 에 반영."
            ),
        },
        "calculators": {},
    }

    ws = find_sheet("잔넬")
    if ws is not None:
        out["calculators"]["channel"] = parse_channel(ws, bcalc["channel"])

    ws = find_sheet("스카시")
    if ws is not None:
        out["calculators"]["gomu"] = parse_gomu(ws, bcalc["gomu"])

    ws = find_sheet("아크릴")
    if ws is not None:
        out["calculators"]["acryl"] = parse_acryl(ws, bcalc["acryl"])

    ws = find_sheet("금은경", "금경", "은경")
    if ws is not None:
        out["calculators"]["goldSilver"] = parse_gold_silver(ws)

    ws = find_sheet("에폭시")
    if ws is not None:
        out["calculators"]["epoxy"] = parse_epoxy(ws, bcalc["epoxy"])

    OUT.write_text(
        json.dumps(out, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # Summary
    c = out["calculators"]
    print(f"OK → {OUT.relative_to(ROOT)}")
    if "channel" in c:
        n = sum(len(t.get("prices") or t.get("pricesByLang", {}).get("eng", {})) +
                (len(t.get("pricesByLang", {}).get("kor", {})) if t["needsLang"] else 0)
                for t in c["channel"]["types"])
        print(f"  channel:    {len(c['channel']['types'])} types, {n} cells")
    if "gomu" in c:
        n = sum(len(d) for d in c["gomu"]["prices"].values())
        print(f"  gomu:       {n} cells")
    if "acryl" in c:
        n = sum(len(d) for tk in c["acryl"]["prices"].values() for d in tk.values())
        print(f"  acryl:      {n} cells")
    if "goldSilver" in c:
        n = sum(len(d) for tk in c["goldSilver"]["prices"].values()
                for tt in tk.values() for d in tt.values())
        print(f"  goldSilver: {n} cells (NEW — Phase 4 에서 확정)")
    if "epoxy" in c:
        n = sum(len(strokes)
                for mat in c["epoxy"]["prices"].values()
                for tt in mat.values()
                for strokes in tt.values())
        print(f"  epoxy:      {n} cells")


if __name__ == "__main__":
    main()
