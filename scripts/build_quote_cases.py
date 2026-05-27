"""
명세서 엑셀 → quote case 그룹화 + 매출 상위 거래처 추출 + vault 페이지 생성.

사용:
    py -3 scripts/build_quote_cases.py dryrun [N]
        매출 상위 N (기본 5) 거래처 + 각 case 그룹 통계
    py -3 scripts/build_quote_cases.py dump-client "<거래처명>"
        해당 거래처 전체 case 그룹 JSON 으로 출력
    py -3 scripts/build_quote_cases.py client-page "<거래처명>"
        해당 거래처 마크다운 페이지를 hdsign-vault/quotes/clients/ 에 작성
    py -3 scripts/build_quote_cases.py category-page "<대분류>"
        해당 카테고리 마크다운 페이지를 hdsign-vault/quotes/categories/ 에 작성
    py -3 scripts/build_quote_cases.py top-clients [N]
        매출 상위 N (기본 5) 거래처 페이지 일괄 생성
    py -3 scripts/build_quote_cases.py all-categories
        모든 자재·작업 카테고리 페이지 일괄 생성
"""

from __future__ import annotations
import json
import re
import statistics
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

NAMESEO_FILES = [
    Path(r"C:\Users\USER\Downloads\files\25년도매출_개인_정리본.xlsx"),
    Path(r"C:\Users\USER\Downloads\files\25년도매출_주식회사_정리본.xlsx"),
]
VAULT = Path(r"C:\Users\USER\Desktop\hdsign-vault")
SHEET = "거래명세"

COLS = {
    "사업자번호": "biz_no",
    "거래처명": "client",
    "발행일자": "issued",
    "월": "month",
    "품목": "item",
    "대분류": "category",
    "합계금액": "total",
    "공급가액": "supply",
    "세액": "tax",
    "면세금액": "tax_free",
    "작성일자": "drafted",
    "발행방법": "method",
    "상태": "status",
}

# 비자재 (잔차 분석 의미 약함) — 통계엔 포함하되 categories 페이지는 따로 표시
NON_MATERIAL = {"할인/DC", "추가/누락/AS", "기타/메모", "이월건", "(미분류)"}


def load_all() -> list[dict]:
    rows: list[dict] = []
    for path in NAMESEO_FILES:
        if not path.is_file():
            print(f"[warn] missing: {path}", file=sys.stderr)
            continue
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[SHEET]
        header: list[str] | None = None
        for raw in ws.iter_rows(values_only=True):
            if header is None:
                header = list(raw)
                continue
            item = {COLS.get(h, h): v for h, v in zip(header, raw)}
            item["_source"] = "개인" if "개인" in path.stem else "주식회사"
            rows.append(item)
    return rows


def normalize_issued(v) -> str:
    return str(v)[:10] if v else ""


def group_by_case(rows: Iterable[dict]) -> dict[tuple, list[dict]]:
    g: dict[tuple, list[dict]] = defaultdict(list)
    for r in rows:
        c = r.get("client")
        i = normalize_issued(r.get("issued"))
        if not c or not i:
            continue
        g[(c, i)].append(r)
    return g


def slug(text: str) -> str:
    s = re.sub(r"[()\[\]{}/\\|:\"<>?*\s]", "", text)
    s = re.sub(r"[\.,;'`~!@#$%^&+=]", "", s)
    return s or "untitled"


def filter_rows(rows: Iterable[dict], client: str | None = None,
                category: str | None = None) -> list[dict]:
    out = []
    for r in rows:
        if client and r.get("client") != client:
            continue
        if category and r.get("category") != category:
            continue
        out.append(r)
    return out


def supply_sum(rows: Iterable[dict]) -> int:
    return sum(int(r.get("supply") or 0) for r in rows)


def total_sum(rows: Iterable[dict]) -> int:
    return sum(int(r.get("total") or 0) for r in rows)


def biz_no_of(rows: Iterable[dict]) -> str:
    for r in rows:
        bn = r.get("biz_no")
        if bn:
            return str(bn)
    return ""


# --- 키워드 마이닝 -----------------------------------------------------------

KEYWORD_PATTERNS = [
    # 두께/사이즈
    (r"(\d+(?:\.\d+)?)\s*[tT]", "두께"),
    (r"(\d+)\s*[xX×]\s*(\d+)", "사이즈"),
    (r"(\d+)\s*[mM][mM]", "사이즈(mm)"),
    # 자재
    (r"갈바", "갈바"),
    (r"스텐", "스텐"),
    (r"포맥스", "포맥스"),
    (r"아크릴", "아크릴"),
    (r"엘이디|LED", "엘이디"),
    (r"잔넬|채널", "잔넬"),
    (r"후렘", "후렘"),
    (r"폴리싱", "폴리싱"),
    (r"도장|페인트", "도장"),
    (r"시트", "시트"),
    (r"실사", "실사"),
    (r"인쇄", "인쇄"),
]


def extract_keywords(items: Iterable[dict]) -> Counter:
    c = Counter()
    for it in items:
        text = str(it.get("item") or "")
        for pat, name in KEYWORD_PATTERNS:
            if re.search(pat, text):
                c[name] += 1
    return c


# --- 마크다운 생성 -----------------------------------------------------------

def fmt_won(v) -> str:
    try:
        return f"₩{int(v):,}"
    except Exception:
        return str(v)


def render_client_page(all_rows: list[dict], client: str) -> str:
    rows = filter_rows(all_rows, client=client)
    if not rows:
        return ""
    biz = biz_no_of(rows)
    source = rows[0].get("_source", "")
    cases = group_by_case(rows)

    # 월별
    by_month = defaultdict(list)
    for r in rows:
        m = r.get("month")
        if m:
            by_month[int(m)].append(r)
    # 카테고리별
    by_cat = defaultdict(list)
    for r in rows:
        by_cat[r.get("category") or "(미분류)"].append(r)
    # 금액 분포
    case_totals = [
        sum(int(i.get("supply") or 0) for i in v)
        for v in cases.values()
    ]
    case_totals = [t for t in case_totals if t > 0]
    if case_totals:
        med = statistics.median(case_totals)
        mn = min(case_totals)
        mx = max(case_totals)
        mean = statistics.mean(case_totals)
    else:
        med = mn = mx = mean = 0

    kws = extract_keywords(rows)

    title = client.replace("/", "／")
    out = []
    out.append("---")
    out.append(f'title: "{title}"')
    out.append(f"type: client")
    out.append(f"tags: [client, 2025]")
    out.append(f"created: 2026-05-27")
    out.append(f"updated: 2026-05-27")
    out.append(f"status: active")
    out.append(f"biz_no: \"{biz}\"")
    out.append(f"source: {source}")
    out.append(f"txn_count: {len(rows)}")
    out.append(f"case_count: {len(cases)}")
    out.append(f"revenue_supply: {supply_sum(rows)}")
    out.append("---")
    out.append("")
    out.append(f"> **{client}** — 사업자번호 `{biz}` ({source}). 2025년 거래 {len(rows)}건, "
               f"케이스(거래일자 기준) {len(cases)}건, 공급가액 누적 **{fmt_won(supply_sum(rows))}**.")
    out.append("")

    # 통계
    out.append("## 통계")
    out.append("")
    out.append("| 항목 | 값 |")
    out.append("|---|---|")
    out.append(f"| 누적 매출 (공급가) | **{fmt_won(supply_sum(rows))}** |")
    out.append(f"| 누적 매출 (합계) | {fmt_won(total_sum(rows))} |")
    out.append(f"| 거래 건수 (행) | {len(rows)} |")
    out.append(f"| 케이스 (거래일자) | {len(cases)} |")
    out.append(f"| 평균 case 금액 | {fmt_won(int(mean))} |")
    out.append(f"| 중앙값 case 금액 | {fmt_won(int(med))} |")
    out.append(f"| 최소 ~ 최대 | {fmt_won(mn)} ~ {fmt_won(mx)} |")
    out.append("")

    # 카테고리 분포
    out.append("## 카테고리 분포")
    out.append("")
    out.append("| 대분류 | 건수 | 공급가 합계 | 비중 |")
    out.append("|---|---|---|---|")
    total_supply_all = supply_sum(rows) or 1
    for cat, items in sorted(by_cat.items(), key=lambda x: -supply_sum(x[1])):
        s = supply_sum(items)
        pct = s * 100.0 / total_supply_all
        out.append(f"| [[quotes/categories/{slug(cat)}|{cat}]] | {len(items)} | {fmt_won(s)} | {pct:.1f}% |")
    out.append("")

    # 월별 추이
    out.append("## 월별 추이")
    out.append("")
    out.append("| 월 | 건수 | 공급가 합계 | 평균 |")
    out.append("|---|---|---|---|")
    for m in sorted(by_month.keys()):
        items = by_month[m]
        s = supply_sum(items)
        avg = s // len(items) if items else 0
        out.append(f"| {m}월 | {len(items)} | {fmt_won(s)} | {fmt_won(avg)} |")
    out.append("")

    # 키워드 마이닝
    if kws:
        out.append("## 자재·공정 키워드 (품목 텍스트 마이닝)")
        out.append("")
        out.append("| 키워드 | 등장 빈도 |")
        out.append("|---|---|")
        for k, v in kws.most_common():
            out.append(f"| {k} | {v} |")
        out.append("")

    # 케이스 표 (대표 30건 — 매출 큰 순)
    out.append("## 케이스 (Top 30, 공급가 큰 순)")
    out.append("")
    out.append("| 발행일자 | 품목 | 대분류 | 공급가 |")
    out.append("|---|---|---|---|")
    case_list = []
    for (c, issued), items in cases.items():
        for it in items:
            case_list.append((issued, it))
    case_list.sort(key=lambda x: -(int(x[1].get("supply") or 0)))
    for issued, it in case_list[:30]:
        item_text = (it.get("item") or "").replace("|", "/")[:50]
        out.append(f"| {issued} | {item_text} | {it.get('category') or '-'} | {fmt_won(it.get('supply'))} |")
    out.append("")

    # 전체 케이스 다운로드 안내
    if len(rows) > 30:
        out.append(f"_({len(rows) - 30}건 더 있음. 전체는 `py scripts/build_quote_cases.py dump-client \"{client}\"` 로 JSON 추출)_")
        out.append("")

    out.append("## 관련")
    out.append("")
    out.append("- [[quotes/README]]")
    out.append("- [[projects/quote-program/index]]")
    return "\n".join(out)


def render_category_page(all_rows: list[dict], category: str) -> str:
    rows = filter_rows(all_rows, category=category)
    if not rows:
        return ""
    by_client = defaultdict(list)
    for r in rows:
        by_client[r.get("client") or "?"].append(r)
    by_month = defaultdict(list)
    for r in rows:
        m = r.get("month")
        if m:
            by_month[int(m)].append(r)
    supplies = [int(r.get("supply") or 0) for r in rows if r.get("supply")]
    if supplies:
        med = statistics.median(supplies)
        mn = min(supplies)
        mx = max(supplies)
        mean = statistics.mean(supplies)
    else:
        med = mn = mx = mean = 0
    kws = extract_keywords(rows)
    is_non_material = category in NON_MATERIAL

    title = category.replace("/", "／")
    out = []
    out.append("---")
    out.append(f'title: "{title}"')
    out.append(f"type: category")
    out.append(f"tags: [category, 2025, {'non_material' if is_non_material else 'material'}]")
    out.append(f"created: 2026-05-27")
    out.append(f"updated: 2026-05-27")
    out.append(f"status: active")
    out.append(f"category: \"{category}\"")
    out.append(f"row_count: {len(rows)}")
    out.append(f"revenue_supply: {supply_sum(rows)}")
    out.append(f"is_non_material: {str(is_non_material).lower()}")
    out.append("---")
    out.append("")
    note = " (비자재 — 가격 분포는 참고)" if is_non_material else ""
    out.append(f"> 대분류 **{category}**{note}. 2025년 {len(rows)}건, "
               f"공급가 누적 **{fmt_won(supply_sum(rows))}**, "
               f"거래처 {len(by_client)}개.")
    out.append("")

    # 통계
    out.append("## 가격 분포")
    out.append("")
    out.append("| 항목 | 값 |")
    out.append("|---|---|")
    out.append(f"| 거래 건수 | {len(rows)} |")
    out.append(f"| 공급가 합계 | **{fmt_won(supply_sum(rows))}** |")
    out.append(f"| 1건 평균 | {fmt_won(int(mean))} |")
    out.append(f"| 1건 중앙값 | {fmt_won(int(med))} |")
    out.append(f"| 1건 최소 ~ 최대 | {fmt_won(mn)} ~ {fmt_won(mx)} |")
    out.append(f"| 거래처 수 | {len(by_client)} |")
    out.append("")

    # 거래처 분포
    out.append("## 거래처 분포 (Top 15)")
    out.append("")
    out.append("| 거래처 | 건수 | 공급가 합계 | 평균 |")
    out.append("|---|---|---|---|")
    sorted_clients = sorted(by_client.items(), key=lambda x: -supply_sum(x[1]))[:15]
    for c, items in sorted_clients:
        s = supply_sum(items)
        avg = s // len(items) if items else 0
        out.append(f"| [[quotes/clients/{slug(c)}|{c}]] | {len(items)} | {fmt_won(s)} | {fmt_won(avg)} |")
    if len(by_client) > 15:
        out.append(f"| _(... {len(by_client) - 15}개 더)_ | | | |")
    out.append("")

    # 월별 추이
    out.append("## 월별 추이")
    out.append("")
    out.append("| 월 | 건수 | 공급가 합계 | 평균 |")
    out.append("|---|---|---|---|")
    for m in sorted(by_month.keys()):
        items = by_month[m]
        s = supply_sum(items)
        avg = s // len(items) if items else 0
        out.append(f"| {m}월 | {len(items)} | {fmt_won(s)} | {fmt_won(avg)} |")
    out.append("")

    # 키워드
    if kws:
        out.append("## 자재·사양 키워드 (품목 텍스트 마이닝)")
        out.append("")
        out.append("| 키워드 | 등장 빈도 |")
        out.append("|---|---|")
        for k, v in kws.most_common():
            out.append(f"| {k} | {v} |")
        out.append("")

    # 상·하위 가격대 sample
    sorted_rows = sorted(rows, key=lambda r: -(int(r.get("supply") or 0)))
    out.append("## 상위 10건 (가격 큰 순)")
    out.append("")
    out.append("| 거래처 | 발행일자 | 품목 | 공급가 |")
    out.append("|---|---|---|---|")
    for r in sorted_rows[:10]:
        item_text = (r.get("item") or "").replace("|", "/")[:40]
        out.append(f"| {r.get('client')} | {normalize_issued(r.get('issued'))} | {item_text} | {fmt_won(r.get('supply'))} |")
    out.append("")
    out.append("## 하위 10건 (가격 작은 순, 0 제외)")
    out.append("")
    out.append("| 거래처 | 발행일자 | 품목 | 공급가 |")
    out.append("|---|---|---|---|")
    low = [r for r in sorted_rows if int(r.get("supply") or 0) > 0][-10:]
    for r in low:
        item_text = (r.get("item") or "").replace("|", "/")[:40]
        out.append(f"| {r.get('client')} | {normalize_issued(r.get('issued'))} | {item_text} | {fmt_won(r.get('supply'))} |")
    out.append("")

    out.append("## 관련")
    out.append("")
    out.append("- [[quotes/README]]")
    out.append("- [[projects/quote-program/index]]")
    return "\n".join(out)


# --- 명령 ---------------------------------------------------------------------

def top_clients_by_revenue(rows: Iterable[dict], n: int) -> list[tuple[str, int, int]]:
    rev: dict[str, int] = defaultdict(int)
    cnt: dict[str, int] = defaultdict(int)
    for r in rows:
        c = r.get("client")
        s = r.get("supply") or 0
        if not c:
            continue
        rev[c] += int(s)
        cnt[c] += 1
    return [
        (name, rev[name], cnt[name])
        for name, _ in sorted(rev.items(), key=lambda x: -x[1])[:n]
    ]


def category_summary(rows: Iterable[dict]) -> dict[str, int]:
    c: dict[str, int] = defaultdict(int)
    for r in rows:
        c[r.get("category") or "(미분류)"] += 1
    return dict(sorted(c.items(), key=lambda x: -x[1]))


def cmd_dryrun(n: int = 5) -> None:
    rows = load_all()
    print(f"== total rows: {len(rows)}")
    cats = category_summary(rows)
    print(f"== category distribution:")
    for k, v in cats.items():
        print(f"  {k:20s}  {v:4d}")
    top = top_clients_by_revenue(rows, n)
    print(f"\n== top {n} clients by revenue:")
    for name, rev, txn in top:
        print(f"  {name:35s}  ₩{rev:>13,}  ({txn} txns)")


def cmd_dump_client(client: str) -> None:
    rows = load_all()
    matched = filter_rows(rows, client=client)
    if not matched:
        print(f"no rows for: {client}", file=sys.stderr)
        sys.exit(1)
    groups = group_by_case(matched)
    cases = []
    for (c, issued), items in sorted(groups.items(), key=lambda x: x[0][1]):
        cases.append(
            {
                "client": c,
                "issued": issued,
                "total_supply": supply_sum(items),
                "total_amount": total_sum(items),
                "items": [
                    {
                        "category": i.get("category"),
                        "item": i.get("item"),
                        "supply": i.get("supply"),
                        "total": i.get("total"),
                    }
                    for i in items
                ],
            }
        )
    print(json.dumps({"client": client, "cases": cases}, ensure_ascii=False, indent=2))


def cmd_client_page(client: str) -> None:
    rows = load_all()
    md = render_client_page(rows, client)
    if not md:
        print(f"no rows for: {client}", file=sys.stderr)
        sys.exit(1)
    out = VAULT / "quotes" / "clients" / f"{slug(client)}.md"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(md, encoding="utf-8")
    print(f"wrote: {out}  ({len(md)} bytes)")


def cmd_category_page(category: str) -> None:
    rows = load_all()
    md = render_category_page(rows, category)
    if not md:
        print(f"no rows for: {category}", file=sys.stderr)
        sys.exit(1)
    out = VAULT / "quotes" / "categories" / f"{slug(category)}.md"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(md, encoding="utf-8")
    print(f"wrote: {out}  ({len(md)} bytes)")


def cmd_top_clients(n: int = 5) -> None:
    rows = load_all()
    top = top_clients_by_revenue(rows, n)
    for name, _, _ in top:
        md = render_client_page(rows, name)
        out = VAULT / "quotes" / "clients" / f"{slug(name)}.md"
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(md, encoding="utf-8")
        print(f"wrote: {out}")


def cmd_all_categories() -> None:
    rows = load_all()
    cats = category_summary(rows)
    for cat in cats:
        if not cat or cat == "(미분류)":
            continue
        md = render_category_page(rows, cat)
        if not md:
            continue
        out = VAULT / "quotes" / "categories" / f"{slug(cat)}.md"
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(md, encoding="utf-8")
        print(f"wrote: {out}")


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(__doc__)
        return 2
    cmd = argv[1]
    if cmd == "dryrun":
        cmd_dryrun(int(argv[2]) if len(argv) > 2 else 5)
    elif cmd == "dump-client":
        cmd_dump_client(argv[2])
    elif cmd == "client-page":
        cmd_client_page(argv[2])
    elif cmd == "category-page":
        cmd_category_page(argv[2])
    elif cmd == "top-clients":
        cmd_top_clients(int(argv[2]) if len(argv) > 2 else 5)
    elif cmd == "all-categories":
        cmd_all_categories()
    else:
        print(f"unknown cmd: {cmd}", file=sys.stderr)
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
