"""홈페이지 거래처 정보 갱신 — 정리본 행에서 phone/email/담당자 추출해서 PUT.

흐름:
  1. picks (전화 매칭/폴백) + 1:1 자동선택 → 거래처별 xls 행 결정
  2. xls 에서 HP1/전화1, 이메일1, 담당자성명1 뽑기
  3. 홈페이지 현재값과 diff → 변경 있는 거래처만 PUT 대상
  4. preview xlsx 저장 + 콘솔 출력. --apply 인자 주면 실제 PUT 실행

사용:
  python tmp_apply_to_homepage.py            # 미리보기만
  python tmp_apply_to_homepage.py --apply    # 실제 갱신
"""
from __future__ import annotations

import json
import re
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path

import xlrd

sys.path.insert(0, r"C:\Users\USER\Desktop\hdsign")
from tmp_filter_xls import keys_for, keys_overlap
from tmp_pick_by_phone import parse_picks, digits

XLS = Path(r"C:\Users\USER\Documents\카카오톡 받은 파일\거래처정보_20260430181444.xls")
CLIENTS_JSON = Path(r"C:\Users\USER\Desktop\hdsign\tmp_clients.json")
PICKS = Path(r"C:\Users\USER\Desktop\hdsign\tmp_picks.txt")
OUT_PREVIEW = Path(r"C:\Users\USER\Desktop\홈페이지갱신_미리보기.xlsx")

API_BASE = "https://hdsign-production.up.railway.app"
ADMIN_USER = "hdno88"
ADMIN_PASS = "hdno0958"


def normalize_phone(s: str) -> str:
    """숫자만 뽑아서 010-XXXX-XXXX / 0XX-XXX-XXXX 형식으로."""
    d = digits(s)
    if not d:
        return ""
    if d.startswith("010") and len(d) == 11:
        return f"{d[:3]}-{d[3:7]}-{d[7:]}"
    if d.startswith("02") and len(d) in (9, 10):
        # 02-XXX-XXXX or 02-XXXX-XXXX
        body = d[2:]
        if len(body) == 7:
            return f"02-{body[:3]}-{body[3:]}"
        return f"02-{body[:4]}-{body[4:]}"
    if len(d) == 10:
        return f"{d[:3]}-{d[3:6]}-{d[6:]}"
    if len(d) == 11:
        return f"{d[:3]}-{d[3:7]}-{d[7:]}"
    return d  # 비정형


def login() -> str:
    body = json.dumps({"username": ADMIN_USER, "password": ADMIN_PASS}).encode("utf-8")
    req = urllib.request.Request(
        f"{API_BASE}/api/auth/login",
        data=body,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=10) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    return data["token"]


def _api_call(token: str, method: str, path: str, payload: dict) -> dict | None:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        f"{API_BASE}{path}",
        data=body,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {token}",
        },
        method=method,
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        msg = e.read().decode("utf-8", errors="replace")
        print(f"[ERROR {e.code}] {method} {path} payload={payload} → {msg[:200]}", file=sys.stderr)
        return None


def put_client(token: str, client_id: int, payload: dict) -> dict | None:
    return _api_call(token, "PUT", f"/api/admin/clients/{client_id}", payload)


def post_client(token: str, payload: dict) -> dict | None:
    return _api_call(token, "POST", "/api/admin/clients", payload)


def main() -> int:
    apply = "--apply" in sys.argv

    book = xlrd.open_workbook(str(XLS))
    sh = book.sheet_by_index(0)
    header = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
    name_col = next(i for i, h in enumerate(header) if "상호" in h)
    tel_col = next(i for i, h in enumerate(header) if "전화1" in h)
    hp_col = next(i for i, h in enumerate(header) if "HP1" in h)
    email_col = next(i for i, h in enumerate(header) if "이메일1" in h)
    charge_col = next(i for i, h in enumerate(header) if "담당자성명1" in h)

    rows = []
    for r in range(1, sh.nrows):
        rows.append(tuple(sh.cell_value(r, c) for c in range(sh.ncols)))

    clients = json.loads(CLIENTS_JSON.read_text(encoding="utf-8"))

    # 거래처별 매칭 xls 행 + 키
    client_rows: dict[int, list[tuple]] = {}
    client_keys_map: dict[int, set[str]] = {}
    for c in clients:
        names = []
        for f in ("companyName", "networkFolderName"):
            v = (c.get(f) or "").strip()
            if v:
                names.append(v)
        ali = (c.get("aliases") or "").strip()
        if ali:
            for tok in re.split(r"[,\s/|]+", ali):
                tok = tok.strip()
                if tok:
                    names.append(tok)
        keys = set()
        for n in names:
            keys |= keys_for(n)
        client_keys_map[c["id"]] = keys
        matched = []
        for row in rows:
            xname = str(row[name_col]).strip()
            if not xname:
                continue
            if keys_overlap(keys, keys_for(xname)):
                matched.append(row)
        client_rows[c["id"]] = matched

    cn_to_clients: dict[str, list[dict]] = {}
    for c in clients:
        for n in (c.get("companyName"), c.get("networkFolderName")):
            if n:
                cn_to_clients.setdefault(n.strip(), []).append(c)

    picks, excludes = parse_picks(PICKS)

    # 제외 거래처 — 정확 이름만 (loose 매칭 금지)
    excluded_ids: set = set()
    for ex in excludes:
        for c in clients:
            cn = (c.get("companyName") or "").strip()
            nf = (c.get("networkFolderName") or "").strip()
            if cn == ex or nf == ex:
                excluded_ids.add(c["id"])

    # 거래처별 → 사용할 xls 행 결정
    chosen: dict[int, tuple] = {}  # client_id → row
    chosen_phone_override: dict[int, str] = {}  # 사용자 입력 전화로 갱신할 거래처
    contacts_by_client: dict[int, list[str]] = {}  # client_id → 담당자명 리스트 (다중 picks 합치기용)
    # 다중 picks 의 추가 담당자 (옵션 C: 신규 거래처로 등록)
    # client_id → list of (phone, contact_name) — 첫 번째는 기존 client UPDATE 용,
    # 두 번째 이후는 신규 PENDING_SIGNUP 거래처 CREATE 용
    multi_picks_by_client: dict[int, list[tuple[str, str]]] = {}

    # picks 와 같은 인덱스의 담당자명
    pick_contacts = getattr(parse_picks, "contacts", [""] * len(picks))

    # 1) picks 처리
    for idx, (name, phone) in enumerate(picks):
        contact_for_pick = pick_contacts[idx] if idx < len(pick_contacts) else ""
        cand: list[dict] = []
        if name in cn_to_clients:
            cand = [c for c in cn_to_clients[name] if c["id"] not in excluded_ids]
        else:
            tk = keys_for(name)
            for c in clients:
                if c["id"] in excluded_ids:
                    continue
                if keys_overlap(client_keys_map[c["id"]], tk):
                    cand.append(c)
        if not cand:
            continue
        # 전화 매칭 우선
        hit = None
        for c in cand:
            for row in client_rows.get(c["id"], []):
                t = digits(str(row[tel_col]))
                h = digits(str(row[hp_col]))
                if (t and t == phone) or (h and h == phone):
                    hit = (c, row)
                    break
            if hit:
                break
        if hit is None:
            # 1:1 폴백 — 단일 행이면 채택, 사용자 전화로 갱신
            single: list[tuple[dict, tuple]] = []
            for c in cand:
                for row in client_rows.get(c["id"], []):
                    single.append((c, row))
            if len(single) == 1:
                hit = single[0]
                if single[0][0]["id"] not in chosen_phone_override:
                    chosen_phone_override[single[0][0]["id"]] = phone
            elif len(single) >= 2:
                tk = keys_for(name)
                def score(r):
                    return -len(tk & keys_for(str(r[name_col]).strip()))
                single.sort(key=lambda cr: score(cr[1]))
                hit = single[0]
                if single[0][0]["id"] not in chosen_phone_override:
                    chosen_phone_override[single[0][0]["id"]] = phone
        if hit is None:
            continue
        cli, row = hit
        # 같은 거래처가 여러 picks 로 들어오면 첫 row 만 chosen 에 (HP1 갱신은 첫 phone)
        # 다중 담당자는 옵션 C: 첫 = 기존 UPDATE, 나머지 = 신규 CREATE.
        if cli["id"] not in chosen:
            chosen[cli["id"]] = row
        if contact_for_pick:
            contacts_by_client.setdefault(cli["id"], []).append(contact_for_pick)
        # 다중 picks 트래킹 — 모든 (phone, contact) 보존
        multi_picks_by_client.setdefault(cli["id"], []).append((phone, contact_for_pick))

    # 2) 1:1 자동선택 — picks 안 됐고 excluded 도 아니고 단일 xls 행 매칭이면 자동
    auto_count = 0
    for c in clients:
        if c["id"] in chosen or c["id"] in excluded_ids:
            continue
        rs = client_rows.get(c["id"], [])
        if len(rs) == 1:
            chosen[c["id"]] = rs[0]
            auto_count += 1

    print(f"[chosen] picks 매칭: {len(chosen) - auto_count}, 1:1 자동선택: {auto_count}, 합계: {len(chosen)}")

    # 3) 거래처별 갱신 계획 — 변경 있는 것만
    # - 단일 pick / 자동선택: UPDATE existing
    # - 다중 picks (옵션 C): 첫 담당자 → UPDATE existing (rename + data), 나머지 → CREATE new
    plans: list[dict] = []
    by_id = {c["id"]: c for c in clients}

    def extract_email(row):
        email_raw = str(row[email_col]).strip()
        m = re.search(r"[\w.+-]+@[\w-]+(?:\.[\w-]+)+", email_raw)
        return m.group(0) if m else ""

    def extract_xls_contact(row):
        c = str(row[charge_col]).strip()
        if c in {"관리자", "담당자", "사장님"}:
            return ""
        if c and re.match(r"^\d{2,3}[\s\-]?\d{3,4}[\s\-]?\d{4}", c):
            return ""
        return c

    for cid, row in chosen.items():
        cli = by_id[cid]
        all_picks = multi_picks_by_client.get(cid, [])
        is_multi = len(all_picks) >= 2

        cur_phone = (cli.get("phone") or "").strip()
        cur_email = (cli.get("email") or "").strip()
        cur_contact = (cli.get("contactName") or "").strip()
        cur_company = (cli.get("companyName") or "").strip()
        new_email = extract_email(row)

        if is_multi:
            # 옵션 C 분기 — 첫 담당자는 UPDATE rename, 나머지는 신규 CREATE
            base_name = cur_company  # 기존 companyName 이 베이스 (예: '진성커뮤니티')
            for idx, (pk_phone, pk_contact) in enumerate(all_picks):
                phone_pretty = normalize_phone(pk_phone)
                contact_short = pk_contact[:50]
                renamed = f"{base_name}({pk_contact})" if pk_contact else base_name
                if idx == 0:
                    # UPDATE existing
                    payload = {}
                    if renamed != cur_company:
                        payload["companyName"] = renamed
                    if phone_pretty and phone_pretty != cur_phone:
                        payload["phone"] = phone_pretty
                    if new_email and new_email.lower() != cur_email.lower():
                        payload["email"] = new_email
                    if contact_short and contact_short != cur_contact:
                        payload["contactName"] = contact_short
                    plans.append({
                        "action": "UPDATE",
                        "id": cid,
                        "companyName": renamed,
                        "xls상호": str(row[name_col]).strip(),
                        "현재phone": cur_phone, "신규phone": phone_pretty,
                        "현재email": cur_email, "신규email": new_email,
                        "현재contact": cur_contact, "신규contact": contact_short,
                        "payload": payload,
                        "변경여부": "변경" if payload else "유지",
                    })
                else:
                    # CREATE new pendingSignup — 이메일은 첫 entry 가 가져갔으니 비움 (unique 제약)
                    payload = {
                        "companyName": renamed,
                        "networkFolderName": "",
                        "phone": phone_pretty,
                        "contactName": contact_short,
                        "pendingSignup": True,
                    }
                    plans.append({
                        "action": "CREATE",
                        "id": None,
                        "companyName": renamed,
                        "xls상호": str(row[name_col]).strip(),
                        "현재phone": "", "신규phone": phone_pretty,
                        "현재email": "", "신규email": "",
                        "현재contact": "", "신규contact": contact_short,
                        "payload": payload,
                        "변경여부": "신규",
                    })
            continue

        # 단일 pick / 자동선택 → 기존 UPDATE 흐름
        if cid in chosen_phone_override:
            new_phone = normalize_phone(chosen_phone_override[cid])
        else:
            hp_raw = digits(str(row[hp_col]))
            tel_raw = digits(str(row[tel_col]))
            new_phone = normalize_phone(hp_raw) if hp_raw else normalize_phone(tel_raw)

        pick_names = contacts_by_client.get(cid, [])
        if pick_names:
            new_contact = (" / ".join(dict.fromkeys(pick_names)))[:50]
        else:
            new_contact = extract_xls_contact(row)

        payload = {}
        if new_phone and new_phone != cur_phone:
            payload["phone"] = new_phone
        if new_email and new_email.lower() != cur_email.lower():
            payload["email"] = new_email
        if new_contact and new_contact != cur_contact:
            payload["contactName"] = new_contact

        plans.append({
            "action": "UPDATE",
            "id": cid,
            "companyName": cur_company,
            "xls상호": str(row[name_col]).strip(),
            "현재phone": cur_phone, "신규phone": new_phone,
            "현재email": cur_email, "신규email": new_email,
            "현재contact": cur_contact, "신규contact": new_contact,
            "payload": payload,
            "변경여부": "변경" if payload else "유지",
        })

    plans.sort(key=lambda p: (p["action"], p["companyName"] or ""))
    n_update = sum(1 for p in plans if p["action"] == "UPDATE" and p["변경여부"] == "변경")
    n_create = sum(1 for p in plans if p["action"] == "CREATE")
    n_keep = sum(1 for p in plans if p["action"] == "UPDATE" and p["변경여부"] == "유지")
    print(f"[plans] 총 {len(plans)}건  UPDATE {n_update}건  CREATE {n_create}건  유지 {n_keep}건")

    # 4) 미리보기 xlsx
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        s = wb.active
        s.title = "갱신계획"
        cols = ["action", "변경여부", "id", "companyName", "xls상호",
                "현재phone", "신규phone", "현재email", "신규email",
                "현재contact", "신규contact", "갱신필드"]
        s.append(cols)
        for c in s[1]:
            c.font = Font(bold=True)
        fill_update = PatternFill("solid", fgColor="FFE5CC")
        fill_create = PatternFill("solid", fgColor="D9F2D9")
        for p in plans:
            s.append([
                p["action"], p["변경여부"], p["id"], p["companyName"], p["xls상호"],
                p["현재phone"], p["신규phone"],
                p["현재email"], p["신규email"],
                p["현재contact"], p["신규contact"],
                ", ".join(p["payload"].keys()),
            ])
            if p["action"] == "CREATE":
                for c in s[s.max_row]:
                    c.fill = fill_create
            elif p["변경여부"] == "변경":
                for c in s[s.max_row]:
                    c.fill = fill_update
        widths = [8, 8, 6, 28, 26, 16, 16, 26, 26, 16, 22, 24]
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, start=1):
            s.column_dimensions[get_column_letter(i)].width = w
        s.freeze_panes = "A2"
        s.auto_filter.ref = s.dimensions
        wb.save(str(OUT_PREVIEW))
        print(f"[saved] {OUT_PREVIEW}")
    except Exception as e:
        print(f"[WARN] 미리보기 저장 실패: {e}", file=sys.stderr)

    if not apply:
        print("\n--apply 없이 실행했으므로 실제 PUT 은 보내지 않았습니다.")
        print("미리보기 확인 후 같은 명령에 --apply 붙여 다시 실행하세요.")
        return 0

    # 5) 실제 API 호출
    print("\n=== APPLY: 홈페이지 PUT/POST 시작 ===")
    token = login()
    ok_u = ok_c = fail = 0
    for p in plans:
        if p["action"] == "UPDATE":
            if p["변경여부"] != "변경":
                continue
            res = put_client(token, p["id"], p["payload"])
            if res is None:
                fail += 1
            else:
                ok_u += 1
        elif p["action"] == "CREATE":
            res = post_client(token, p["payload"])
            if res is None:
                fail += 1
            else:
                ok_c += 1
        time.sleep(0.05)
    print(f"[apply] UPDATE 성공 {ok_u}  CREATE 성공 {ok_c}  실패 {fail}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
