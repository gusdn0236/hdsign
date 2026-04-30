from openpyxl import load_workbook
wb = load_workbook(r"C:\Users\USER\Desktop\홈페이지갱신_미리보기.xlsx")
s = wb.active
rows = list(s.iter_rows(values_only=True))
# 컬럼: action, 변경여부, id, companyName, xls상호, 현재phone, 신규phone, 현재email, 신규email, 현재contact, 신규contact, 갱신필드
print("진성커뮤니티 / 디온에이 처리:\n")
for r in rows[1:]:
    cn = (r[3] or "")
    if "진성커뮤니티" in cn or "디온에이" in cn:
        action = r[0]
        print(f"  [{action}] id={r[2]}  '{cn}'")
        print(f"    phone   → {r[6]}")
        print(f"    email   → {r[8]}")
        print(f"    contact → {r[10]}")
        print()
