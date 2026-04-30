"""중복 매칭된 거래처에서, 사용자가 전화번호로 지정한 진짜 거래처 행만 남긴다.

입력: tmp_picks.txt — '거래처명 전화번호' (한 줄에 하나)
- 거래처명: 홈페이지 companyName 와 keys_for() 정규화로 매칭
- 전화번호: 숫자만 추출해 비교 (xls 의 전화1 + HP1 둘 다 후보)

출력:
- 거래처_정리본.xlsx (선택된 행만 — 거래처별 1행)
- 콘솔에 미매칭/모호 사례 보고
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import xlrd

sys.path.insert(0, r"C:\Users\USER\Desktop\hdsign")
from tmp_filter_xls import keys_for, keys_overlap  # 동일 정규화 재사용

XLS = Path(r"C:\Users\USER\Documents\카카오톡 받은 파일\거래처정보_20260430181444.xls")
CLIENTS_JSON = Path(r"C:\Users\USER\Desktop\hdsign\tmp_clients.json")
PICKS = Path(r"C:\Users\USER\Desktop\hdsign\tmp_picks.txt")
OUT = Path(r"C:\Users\USER\Desktop\거래처_정리본.xlsx")


def digits(s: str) -> str:
    return re.sub(r"\D", "", s or "")


def parse_picks(path: Path) -> tuple[list[tuple[str, str]], list[str]]:
    """각 줄:
    - '거래처명 전화번호'             → pick (contactName 없음)
    - '거래처명 전화번호 담당자명'    → pick + contactName
    - '!거래처명'                     → 최종 결과에서 제외
    - '#' 시작                       → 주석

    호환성 유지를 위해 picks 는 (name, phone) 튜플로 반환. 담당자명은 별도 모듈 attribute 로.
    """
    picks: list[tuple[str, str]] = []
    excludes: list[str] = []
    contacts: list[str] = []  # picks 와 같은 인덱스로 매핑되는 담당자명 (없으면 "")
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("!"):
            ex = line[1:].strip()
            if ex:
                excludes.append(ex)
            continue
        # 거래처명 + 전화번호 + (선택) 담당자명
        m = re.match(r"^(.+?)\s+([0-9][0-9\-\s]{7,})\s*(\S.*)?$", line)
        if not m:
            print(f"[skip] 파싱 실패: {line}")
            continue
        name = m.group(1).strip()
        phone = digits(m.group(2))
        contact = (m.group(3) or "").strip()
        if len(phone) < 9:
            print(f"[skip] 전화 너무 짧음: {line}")
            continue
        picks.append((name, phone))
        contacts.append(contact)
    parse_picks.contacts = contacts  # type: ignore[attr-defined]
    return picks, excludes


def main() -> int:
    book = xlrd.open_workbook(str(XLS))
    sh = book.sheet_by_index(0)
    header = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
    name_col = next(i for i, h in enumerate(header) if "상호" in h)
    tel_col = next(i for i, h in enumerate(header) if "전화1" in h)
    hp_col = next(i for i, h in enumerate(header) if "HP1" in h)

    rows = []
    for r in range(1, sh.nrows):
        rows.append(tuple(sh.cell_value(r, c) for c in range(sh.ncols)))

    clients = json.loads(CLIENTS_JSON.read_text(encoding="utf-8"))

    # 거래처별 매칭 xls 행 (전체 — 중복 포함)
    client_rows: dict[int, list[tuple]] = {}
    client_keys: dict[int, set[str]] = {}
    for c in clients:
        names = []
        for f in ("companyName", "networkFolderName"):
            v = (c.get(f) or "").strip()
            if v:
                names.append(v)
        ali = (c.get("aliases") or "").strip()
        if ali:
            for tok in re.split(r"[,\s/|]+", ali):
                tok = tok.strip()
                if tok:
                    names.append(tok)
        keys = set()
        for n in names:
            keys |= keys_for(n)
        client_keys[c["id"]] = keys
        matched = []
        for row in rows:
            xname = str(row[name_col]).strip()
            if not xname:
                continue
            if keys_overlap(keys, keys_for(xname)):
                matched.append(row)
        client_rows[c["id"]] = matched

    # cn → client id (중복 가능 — keys_for 동일하면 여러 후보)
    cn_to_clients: dict[str, list[dict]] = {}
    for c in clients:
        for n in (c.get("companyName"), c.get("networkFolderName")):
            if not n:
                continue
            n = n.strip()
            cn_to_clients.setdefault(n, []).append(c)

    picks, excludes = parse_picks(PICKS)
    print(f"[picks] 입력 {len(picks)}건  제외 {len(excludes)}건")

    # 제외 폴더 → client_id 화 (companyName/networkFolderName 정확/정규화 일치)
    excluded_client_ids: set = set()
    for ex in excludes:
        ex_keys = keys_for(ex)
        hit = False
        for c in clients:
            cn = (c.get("companyName") or "").strip()
            nf = (c.get("networkFolderName") or "").strip()
            if cn == ex or nf == ex:
                excluded_client_ids.add(c["id"])
                hit = True
        if hit:
            continue
        # 키 정규화 매칭 폴백
        for c in clients:
            ck = client_keys.get(c["id"], set())
            if keys_overlap(ck, ex_keys):
                excluded_client_ids.add(c["id"])
                hit = True
        if not hit:
            print(f"[제외-스킵] 홈페이지 거래처 매칭 없음: {ex}")

    selected: list[tuple[dict, tuple, str]] = []  # (client, row, phone입력)
    not_found_client: list[tuple[str, str]] = []
    not_found_phone: list[tuple[str, str, list[str]]] = []  # name, phone, 후보전화들
    ambiguous: list[tuple[str, str, list[str]]] = []
    fallback_picked: list[tuple[str, str, str]] = []  # name, phone입력, xls상호 — 전화 불일치인데 1:1이라 채택

    for name, phone in picks:
        # 1) 거래처 찾기 — 정확 일치 우선.
        # 정확 일치가 있는데 제외돼서 cand_clients 가 비면 폴백하지 않는다 (다른 거래처로 우회 금지).
        cand_clients: list[dict] = []
        if name in cn_to_clients:
            cand_clients = [c for c in cn_to_clients[name] if c["id"] not in excluded_client_ids]
        else:
            target_keys = keys_for(name)
            for c in clients:
                if c["id"] in excluded_client_ids:
                    continue
                if keys_overlap(client_keys[c["id"]], target_keys):
                    cand_clients.append(c)

        if not cand_clients:
            not_found_client.append((name, phone))
            continue

        # 2) 후보 거래처들의 xls 행 중 전화 일치 행 찾기
        hit_rows: list[tuple[dict, tuple]] = []
        all_phones_seen: list[str] = []
        for c in cand_clients:
            for row in client_rows.get(c["id"], []):
                ph_tel = digits(str(row[tel_col]))
                ph_hp = digits(str(row[hp_col]))
                if ph_tel:
                    all_phones_seen.append(ph_tel)
                if ph_hp:
                    all_phones_seen.append(ph_hp)
                if (ph_tel and ph_tel == phone) or (ph_hp and ph_hp == phone):
                    hit_rows.append((c, row))

        if not hit_rows:
            # 전화 불일치 폴백 — 후보 거래처가 명확하면(1:1 또는 이름 베스트매치) 행 채택,
            # 사용자 입력 전화로 갱신해서 출력. 사용자가 explicit pick 으로 알려줬으니 행 자체는 맞다고 가정.
            fb_row = None
            fb_client = None
            # 1) 단일 후보
            single_rows: list[tuple[dict, tuple]] = []
            for c in cand_clients:
                for row in client_rows.get(c["id"], []):
                    single_rows.append((c, row))
            if len(single_rows) == 1:
                fb_client, fb_row = single_rows[0]
            elif len(single_rows) >= 2:
                # 이름 거리 — 입력 거래처와 가장 가까운 xls 상호
                target_keys = keys_for(name)
                def score(r):
                    xname = str(r[name_col]).strip()
                    xks = keys_for(xname)
                    inter = len(target_keys & xks)
                    return (-inter, len(xname))
                single_rows.sort(key=lambda cr: score(cr[1]))
                fb_client, fb_row = single_rows[0]

            if fb_row is not None:
                # 사용자 전화로 HP1 덮어쓰기 — 출력 정합성 위해 row 를 mutable list 로
                row_list = list(fb_row)
                pretty = phone if len(phone) < 4 else f"{phone[:3]}-{phone[3:-4]}-{phone[-4:]}"
                if hp_col >= 0 and hp_col < len(row_list):
                    row_list[hp_col] = pretty
                selected.append((fb_client, tuple(row_list), phone))
                fallback_picked.append((name, phone, str(fb_row[name_col]).strip()))
                continue
            not_found_phone.append((name, phone, all_phones_seen))
            continue

        if len(hit_rows) > 1:
            # 같은 전화가 여러 행에 — 흔치 않으나 보고
            descs = [str(r[name_col]) for _, r in hit_rows]
            ambiguous.append((name, phone, descs))

        c, row = hit_rows[0]
        selected.append((c, row, phone))

    # ── 출력 ─────────────────────────────
    print()
    print(f"[result] 선택 성공: {len(selected)}")
    print(f"[result] 거래처 못 찾음: {len(not_found_client)}")
    print(f"[result] 전화 못 찾음: {len(not_found_phone)}")
    print(f"[result] 동일 전화 다중: {len(ambiguous)}")
    print(f"[result] 폴백 선택(전화 불일치 → 1:1/이름매칭으로 채택, HP1 사용자전화로 갱신): {len(fallback_picked)}")

    if not_found_client:
        print("\n[못 찾음] 거래처명 매칭 실패:")
        for n, p in not_found_client:
            print(f"  - {n}  ({p})")
    if not_found_phone:
        print("\n[못 찾음] 거래처는 매칭됐으나 전화 일치 행 없음:")
        for n, p, seen in not_found_phone:
            short = ", ".join(sorted(set(seen))[:6])
            print(f"  - {n}  입력전화={p}  xls전화후보=[{short}]")
    if ambiguous:
        print("\n[모호] 동일 전화가 xls 여러 행에 — 첫 행 선택:")
        for n, p, ds in ambiguous:
            print(f"  - {n}  ({p}) → {ds}")
    if fallback_picked:
        print("\n[폴백] 전화 불일치 → 1:1/이름매칭 행 채택, HP1 을 사용자전화로 갱신:")
        for n, p, xn in fallback_picked:
            print(f"  - {n}  (입력{p}) → xls:{xn}")

    # xlsx 저장
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        wb = Workbook()
        s = wb.active
        s.title = "거래처(정리본)"
        s.append(["거래처(홈페이지)", *header])
        for c in s[1]:
            c.font = Font(bold=True)
        for c, row, _ in selected:
            cn = c.get("companyName") or ""
            cells = []
            for v in row:
                if isinstance(v, float) and v.is_integer():
                    cells.append(int(v))
                else:
                    cells.append(v)
            s.append([cn, *cells])
        s.freeze_panes = "A2"

        # 남은 거래처 — picks 안 됐고 제외도 아닌 거래처 (입력 진행도 추적용)
        picked_ids = {c["id"] for c, _, _ in selected}
        remaining = [
            c for c in clients
            if c["id"] not in picked_ids and c["id"] not in excluded_client_ids
        ]
        remaining.sort(key=lambda x: (x.get("companyName") or "").strip())
        s3 = wb.create_sheet("전화입력필요")
        s3.append(["companyName", "networkFolderName", "aliases", "status", "id"])
        for c in s3[1]:
            c.font = Font(bold=True)
        for c in remaining:
            s3.append([
                c.get("companyName") or "",
                c.get("networkFolderName") or "",
                c.get("aliases") or "",
                c.get("status") or "",
                c.get("id"),
            ])
        s3.column_dimensions["A"].width = 28
        s3.column_dimensions["B"].width = 22
        s3.freeze_panes = "A2"

        # 미매칭 시트 — 추가 입력 필요한 항목
        if not_found_client or not_found_phone:
            s2 = wb.create_sheet("미매칭_재확인필요")
            s2.append(["입력거래처", "입력전화", "사유", "xls전화후보(최대 6개)"])
            for n, p in not_found_client:
                s2.append([n, p, "거래처명 매칭 실패", ""])
            for n, p, seen in not_found_phone:
                short = ", ".join(sorted(set(seen))[:6])
                s2.append([n, p, "거래처 OK / 전화 불일치", short])

        wb.save(str(OUT))
        print(f"\n[saved] {OUT}")
        print(f"        - 정리본: {len(selected)}행")
        print(f"        - 전화입력필요: {len(remaining)}개 거래처")
        print(f"        - 제외 적용: {len(excluded_client_ids)}개 거래처")
    except Exception as e:
        print(f"[WARN] 저장 실패: {e}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
