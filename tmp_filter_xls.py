"""거래처정보_*.xls 를 홈페이지 거래처 목록과 매칭해 필터링.

기준: 홈페이지(/api/admin/clients)에 등록된 거래처. 네트워크 폴더는 사용하지 않음.
거래처 단위로 매칭 키 묶음을 만들고, xls 의 각 행을 그 키 묶음과 비교.

매칭 키 생성:
- companyName / networkFolderName / aliases(콤마구분) 모든 이름에 대해 keys_for() 적용
- keys_for: 회사형태 제거 + 괄호 안/밖 분리 + 후행 일반명사(광고/디자인/사인 등) 제거 → 다중 키 생성
- 매칭: 키셋 교집합, 또는 길이 3 이상 부분문자열 일치
"""
from __future__ import annotations

import json
import re
import sys
import unicodedata
from pathlib import Path

import xlrd

XLS = Path(r"C:\Users\USER\Documents\카카오톡 받은 파일\거래처정보_20260430181444.xls")
CLIENTS_JSON = Path(r"C:\Users\USER\Desktop\hdsign\tmp_clients.json")
OUT_FILTERED = Path(r"C:\Users\USER\Desktop\거래처_필터결과.xlsx")
OUT_MISSING = Path(r"C:\Users\USER\Desktop\누락거래처.xlsx")
OUT_DUPS = Path(r"C:\Users\USER\Desktop\중복거래처.xlsx")

_CORP_FORMS = [
    "주식회사", "유한회사", "유한책임회사", "합자회사", "합명회사",
    "(주)", "(유)", "㈜", "㈲",
]

_PAREN_RE = re.compile(r"[\(（].*?[\)）]")
_PAREN_CONTENT_RE = re.compile(r"[\(（]([^()（）]+)[\)）]")
_PREFIX_NOISE_RE = re.compile(r"^[*\s]+|^폐업[\s\-]*")

_TAIL_TOKENS = [
    "종합광고기획", "종합광고", "광고기획", "디자인기획",
    "광고", "사인", "디자인", "기획", "산업", "이엔지", "엔지니어링",
    "테크", "에이엔디", "애드", "ad", "design", "sign",
    "주식회사", "코퍼레이션", "엔터프라이즈",
]

# 괄호 안 내용 중 매칭 키로 쓰면 안 되는 것 — 지명/직책 등.
_REGION_OR_NOISE = {
    "인천", "강남", "전주", "전북", "광명", "성남", "대전", "대구", "부산",
    "제주", "창원", "양주", "서울", "울산", "수원", "고양", "용인",
    "안양", "안산", "분당", "일산", "김포", "파주", "포천", "청주",
    "안양점", "강북", "강서", "강동",
}


def norm_key(s: str) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFC", str(s)).strip()
    for form in _CORP_FORMS:
        s = s.replace(form, "")
    s = _PAREN_RE.sub("", s)
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[.\-_/]", "", s)
    return s.lower()


def core_key(k: str) -> str:
    if not k:
        return k
    out = k
    changed = True
    while changed:
        changed = False
        for t in _TAIL_TOKENS:
            if out.endswith(t) and len(out) - len(t) >= 2:
                out = out[: -len(t)]
                changed = True
                break
    return out or k


def keys_for(name: str) -> set[str]:
    """이름 하나에서 매칭 후보가 될 정규화 키 묶음."""
    if not name:
        return set()
    s = unicodedata.normalize("NFC", str(name)).strip()
    s = _PREFIX_NOISE_RE.sub("", s)
    for form in _CORP_FORMS:
        s = s.replace(form, "")
    out: set[str] = set()
    parens = _PAREN_CONTENT_RE.findall(s)
    main = _PAREN_CONTENT_RE.sub("", s)
    main_n = norm_key(main)
    if main_n:
        out.add(main_n)
        ck = core_key(main_n)
        if ck and ck != main_n:
            out.add(ck)
    full_n = norm_key(s)
    if full_n:
        out.add(full_n)
    for p in parens:
        p = p.strip()
        if not p or p.endswith("님") or p in _REGION_OR_NOISE:
            continue
        pn = norm_key(p)
        if pn and len(pn) >= 2:
            out.add(pn)
            cpn = core_key(pn)
            if cpn and cpn != pn and len(cpn) >= 2:
                out.add(cpn)
    return {k for k in out if k}


def keys_overlap(a: set[str], b: set[str]) -> bool:
    if not a or not b:
        return False
    if a & b:
        return True
    for fk in a:
        if len(fk) < 2:
            continue
        for xk in b:
            if len(xk) < 2:
                continue
            if fk == xk:
                return True
            if len(fk) >= 3 or len(xk) >= 3:
                if fk in xk or xk in fk:
                    return True
    return False


def find_company_col(sheet) -> int:
    header = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
    for c, name in enumerate(header):
        if any(k in name for k in ("상호", "거래처명", "회사명", "업체명", "거래처")):
            return c
    return 0


def main() -> int:
    if not XLS.exists():
        print(f"[ERROR] xls 없음: {XLS}", file=sys.stderr)
        return 2
    if not CLIENTS_JSON.exists():
        print(f"[ERROR] 거래처 JSON 없음: {CLIENTS_JSON}", file=sys.stderr)
        return 2

    book = xlrd.open_workbook(str(XLS))
    sheet = book.sheet_by_index(0)
    print(f"[xls] sheet={sheet.name!r}  rows={sheet.nrows}  cols={sheet.ncols}")
    header = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
    company_col = find_company_col(sheet)
    print(f"[xls] 상호 컬럼: {company_col} ({header[company_col]!r})")

    clients = json.loads(CLIENTS_JSON.read_text(encoding="utf-8"))
    print(f"[clients] 홈페이지 거래처 수: {len(clients)}")

    # 거래처마다 키 묶음과 모든 별칭 보존.
    client_entries: list[dict] = []
    for c in clients:
        names: list[str] = []
        for f in ("companyName", "networkFolderName"):
            v = (c.get(f) or "").strip()
            if v:
                names.append(v)
        ali = (c.get("aliases") or "").strip()
        if ali:
            for tok in re.split(r"[,\s/|]+", ali):
                tok = tok.strip()
                if tok:
                    names.append(tok)
        keys: set[str] = set()
        for n in names:
            keys |= keys_for(n)
        client_entries.append({
            "id": c.get("id"),
            "companyName": c.get("companyName"),
            "networkFolderName": c.get("networkFolderName"),
            "aliases": c.get("aliases"),
            "status": c.get("status"),
            "names": names,
            "keys": keys,
        })

    rows_all: list[tuple] = []
    for r in range(1, sheet.nrows):
        rows_all.append(tuple(sheet.cell_value(r, c) for c in range(sheet.ncols)))

    xls_keysets: list[tuple[tuple, str, set[str]]] = []
    for row in rows_all:
        name = str(row[company_col]).strip()
        if not name:
            continue
        xls_keysets.append((row, name, keys_for(name)))

    # 거래처 관점 매칭 — 각 거래처에 대해 매칭되는 모든 xls 행을 찾음.
    # (xls 행 관점이면 한 행이 한 거래처에만 귀속되어, 키가 겹치는 다른 거래처는 누락 처리됨)
    matched_xls_row_ids: set[int] = set()
    matched_client_ids: set = set()
    client_to_rows: dict[int, list[tuple]] = {}

    for ce in client_entries:
        for row, name, xks in xls_keysets:
            if keys_overlap(ce["keys"], xks):
                client_to_rows.setdefault(ce["id"], []).append(row)
                matched_xls_row_ids.add(id(row))
                matched_client_ids.add(ce["id"])

    # 매칭된 행은 한 번만 — 거래처 여러 개에 매칭됐어도 dedup
    matched_rows: list[tuple] = []
    seen_row_ids: set[int] = set()
    for row, _, _ in xls_keysets:
        if id(row) in matched_xls_row_ids and id(row) not in seen_row_ids:
            matched_rows.append(row)
            seen_row_ids.add(id(row))

    unmatched_xls = [(name, row) for row, name, _ in xls_keysets if id(row) not in matched_xls_row_ids]
    unmatched_clients = [ce for ce in client_entries if ce["id"] not in matched_client_ids]

    print()
    print(f"[result] xls 행수: {len(rows_all)}")
    print(f"[result] 매칭된 xls 행: {len(matched_rows)}")
    print(f"[result] 매칭된 거래처(unique): {len(matched_client_ids)}")
    print(f"[result] 매칭 안 된 거래처: {len(unmatched_clients)}")
    print(f"[result] xls 에 있으나 거래처 매칭 안 된 행: {len(unmatched_xls)}")

    # ── 결과1: 필터된 xls (홈페이지 거래처에 매칭된 행만)
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "거래처(필터됨)"
        ws.append(header)
        for row in matched_rows:
            ws.append(list(row))
        ws2 = wb.create_sheet("거래처만_xls없음")
        ws2.append(["거래처ID", "companyName", "networkFolderName", "aliases", "status"])
        for ce in unmatched_clients:
            ws2.append([
                ce["id"], ce["companyName"], ce["networkFolderName"],
                ce["aliases"], ce["status"],
            ])
        wb.save(str(OUT_FILTERED))
        print(f"\n[saved] {OUT_FILTERED}")
    except Exception as e:
        print(f"[WARN] 필터 결과 xlsx 저장 실패: {e}", file=sys.stderr)

    # ── 결과2: 누락거래처 (홈페이지엔 있는데 xls 에 없는 거래처) — 사용자 요청
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        wb2 = Workbook()
        s1 = wb2.active
        s1.title = "누락거래처"
        s1.append(["companyName", "networkFolderName", "aliases", "status", "id"])
        for c in s1[1]:
            c.font = Font(bold=True)
        # 보기 좋게 companyName 가나다순
        unmatched_clients_sorted = sorted(
            unmatched_clients,
            key=lambda x: (x.get("companyName") or "").strip(),
        )
        for ce in unmatched_clients_sorted:
            s1.append([
                ce["companyName"], ce["networkFolderName"],
                ce["aliases"], ce["status"], ce["id"],
            ])
        s1.column_dimensions["A"].width = 28
        s1.column_dimensions["B"].width = 22
        s1.column_dimensions["C"].width = 22
        s1.freeze_panes = "A2"
        wb2.save(str(OUT_MISSING))
        print(f"[saved] {OUT_MISSING}  ({len(unmatched_clients)}행)")
    except Exception as e:
        print(f"[WARN] 누락거래처 xlsx 저장 실패: {e}", file=sys.stderr)

    # ── 결과3: 중복매칭 (한 거래처가 xls 의 여러 행과 매칭) — 사용자 요청
    dup_clients = {cid: rows for cid, rows in client_to_rows.items() if len(rows) >= 2}
    dup_total = sum(len(v) for v in dup_clients.values())

    def _idx(*candidates: str) -> int:
        for i, h in enumerate(header):
            for c in candidates:
                if c in h:
                    return i
        return -1

    i_corp = _idx("회사구분")
    i_biz = _idx("사업자번호")
    i_name = company_col
    i_owner = _idx("성명")
    i_addr = _idx("주소")
    i_tel = _idx("전화1")
    i_hp = _idx("HP1")
    i_dept = _idx("담당부서1")
    i_charge = _idx("담당자성명1")
    i_email = _idx("이메일1")

    def _cell(row: tuple, idx: int) -> str:
        if idx < 0 or idx >= len(row):
            return ""
        v = row[idx]
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb3 = Workbook()
        s = wb3.active
        s.title = "중복거래처(고르기)"
        s.append([
            "선택", "거래처(홈페이지)", "매칭수",
            "회사구분", "사업자번호", "xls상호", "대표자",
            "주소", "전화1", "HP1", "담당부서1", "담당자1", "이메일1",
        ])
        for c in s[1]:
            c.font = Font(bold=True)

        fill_a = PatternFill("solid", fgColor="FFF7E6")
        fill_b = PatternFill("solid", fgColor="FFFFFFFF")

        # client_id → companyName 캐시
        cid_to_name = {ce["id"]: ce["companyName"] for ce in client_entries}
        ordered = sorted(
            dup_clients.items(),
            key=lambda x: (-len(x[1]), cid_to_name.get(x[0], "")),
        )
        toggle = False
        for cid, rows in ordered:
            toggle = not toggle
            fill = fill_a if toggle else fill_b
            cname = cid_to_name.get(cid, str(cid))
            for row in rows:
                s.append([
                    "", cname, len(rows),
                    _cell(row, i_corp), _cell(row, i_biz), _cell(row, i_name),
                    _cell(row, i_owner), _cell(row, i_addr),
                    _cell(row, i_tel), _cell(row, i_hp),
                    _cell(row, i_dept), _cell(row, i_charge), _cell(row, i_email),
                ])
                for c in s[s.max_row]:
                    c.fill = fill

        widths = [6, 26, 8, 12, 16, 32, 12, 50, 16, 16, 14, 14, 28]
        for i, w in enumerate(widths, start=1):
            from openpyxl.utils import get_column_letter
            s.column_dimensions[get_column_letter(i)].width = w
        s.freeze_panes = "A2"
        s.auto_filter.ref = s.dimensions

        wb3.save(str(OUT_DUPS))
        print(f"[saved] {OUT_DUPS}  ({len(dup_clients)}거래처/{dup_total}행)")
    except Exception as e:
        print(f"[WARN] 중복거래처 xlsx 저장 실패: {e}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
